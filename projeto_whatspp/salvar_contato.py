import openpyxl
from openpyxl import Workbook, load_workbook
import os

# Caminho do arquivo Excel
dir_path = "projeto_whatspp"
file_path = os.path.join(dir_path, "historico_conversas.xlsx")

def inicializar_arquivo_excel():
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversas"
        ws.append(["Contato", "Historico"])
        wb.save(file_path)

def salvar_historico(contato, historico_conversa):
    wb = load_workbook(file_path)
    ws = wb["Conversas"]

    historico_str = "\n".join(historico_conversa)
    encontrado = False

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == contato:
            row[1].value = historico_str
            encontrado = True
            break

    if not encontrado:
        ws.append([contato, historico_str])

    wb.save(file_path)

def carregar_historico(contato):
    wb = load_workbook(file_path)
    ws = wb["Conversas"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == contato:
            return row[1].splitlines()

    return []

def salvar_ultima_mensagem(mensagem):
    wb = load_workbook(file_path)
    if "UltimaMensagem" not in wb.sheetnames:
        ws = wb.create_sheet("UltimaMensagem")
        ws.append(["UltimaMensagem"])
    else:
        ws = wb["UltimaMensagem"]

    if ws.max_row > 1:
        ws.cell(row=2, column=1, value=mensagem)
    else:
        ws.append([mensagem])

    wb.save(file_path)

def carregar_ultima_mensagem():
    wb = load_workbook(file_path)
    if "UltimaMensagem" in wb.sheetnames:
        ws = wb["UltimaMensagem"]
        if ws.max_row > 1:
            return ws.cell(row=2, column=1).value

    return ""
